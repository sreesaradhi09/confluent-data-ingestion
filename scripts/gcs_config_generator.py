#!/usr/bin/env python3
"""
gcs_config_generator.py

Single-file utility to:
1) Create an Excel template for GCS Source/Sink connector inputs
2) Read a filled template and generate deployment-agnostic CONFIG JSON files

Usage:
  # Create template
  python gcs_config_generator.py template --out templates/gcs_connectors_template.xlsx

  # Generate config JSONs
  python gcs_config_generator.py generate --excel templates/gcs_connectors_template.xlsx --outdir out/configs
"""

import argparse
import json
import re
from pathlib import Path
from typing import Any, Dict, List, Tuple, Optional, Literal

import pandas as pd
import pytz
from pydantic import BaseModel, Field, field_validator, ValidationError

# -------------------
# Key normalization & sensitive keys
# -------------------

SENSITIVE_KEYS = {
    "gcs.credentials.config",
    "gcs.credentials.json",
    "basic.auth.user.info",
    "sasl.jaas.config",
    "kafka.api.key",
    "kafka.api.secret",
}

NORMALIZE_MAP = {
    "cloud provider": "cloud.provider",
    "cloud. environment": "cloud.environment",
    "connector. class": "connector.class",
    "gcs. bucket.name": "gcs.bucket.name",
    "gcs. bucket name": "gcs.bucket.name",
    "gcs credentials.json": "gcs.credentials.json",
    "input data format": "input.data.format",
    "output data. format": "output.data.format",
    "output data.format": "output.data.format",
    "input.data. format": "input.data.format",
    "tasks-max": "tasks.max",
    "topic-regex. list": "topic-regex.list",
    "value. converter. decimal. format": "value.converter.decimal.format",
    "value.converter.replace.null.with.default": "value.converter.replace.null.with.default",
    "gcs. bucket name": "gcs.bucket.name",
}

def normalize_key(key: str) -> str:
    k = key.strip()
    k = re.sub(r"\s*\.\s*", ".", k)  # collapse spaces around dots
    k = re.sub(r"\s{2,}", " ", k)    # collapse inner multiple spaces
    return NORMALIZE_MAP.get(k, k)

# Boolean-like fields that should normalize to 'true'/'false' strings
BOOLISH_KEYS = {
    "errors.log.enable",
    "errors.deadletterqueue.context.headers.enable",
    "value.converter.replace.null.with.default",
}

def _coerce_boolish(val: str) -> str:
    r"""Normalize common truthy/falsey inputs to 'true' or 'false' (strings)."""
    if val is None:
        return val
    s = str(val).strip().lower()
    if s in {"true","t","yes","y","1"}:
        return "true"
    if s in {"false","f","no","n","0"}:
        return "false"
    return val  # leave as-is; pydantic will catch invalids

def split_sensitive(config: Dict[str, Any]) -> Tuple[Dict[str, Any], Dict[str, Any]]:
    nonsens, sens = {}, {}
    for k, v in config.items():
        if k in SENSITIVE_KEYS:
            sens[k] = v
        else:
            nonsens[k] = v
    return nonsens, sens

# -------------------
# Pydantic models for validation
# -------------------

AllowedStatus = Literal["RUNNING","PAUSED"]
AllowedEnv = Literal["dev","test","stage","prod"]
AllowedFormats = Literal["JSON","AVRO","PARQUET","CSV","STRING"]
AllowedTimeInterval = Literal["MINUTE","HOURLY","DAILY"]

class CommonRow(BaseModel):
    connector_type: Literal["sink","source"]
    name: str
    cloud_environment: AllowedEnv = Field(validation_alias="cloud.environment")
    kafka_endpoint: str = Field(validation_alias="kafka.endpoint")
    kafka_region: str = Field(validation_alias="kafka.region")
    kafka_service_account_id: str = Field(validation_alias="kafka.service.account.id")
    topics_dir: str = Field(validation_alias="topics.dir")
    tasks_max: str = Field(validation_alias="tasks.max")
    status: AllowedStatus = "PAUSED"

    @field_validator("kafka_endpoint")
    @classmethod
    def endpoint_format(cls, v: str) -> str:
        if "://" not in v or not v.endswith(":9092"):
            raise ValueError("kafka.endpoint should include scheme (e.g., SASL_SSL://) and end with :9092")
        return v

    @field_validator("kafka_region")
    @classmethod
    def region_nonempty(cls, v: str) -> str:
        if not v:
            raise ValueError("kafka.region is required")
        return v

    @field_validator("topics_dir")
    @classmethod
    def topics_dir_not_empty(cls, v: str) -> str:
        if not v:
            raise ValueError("topics.dir is required")
        return v

class SinkExtras(BaseModel):
    topics: str
    gcs_bucket_name: str = Field(validation_alias="gcs.bucket.name")
    input_data_format: AllowedFormats = Field(validation_alias="input.data.format")
    output_data_format: AllowedFormats = Field(validation_alias="output.data.format")
    time_interval: Optional[AllowedTimeInterval] = Field(default=None, validation_alias="time.interval")
    timezone: Optional[str] = None
    transforms: Optional[str] = None
    transforms_MaskField_fields: Optional[str] = Field(default=None, validation_alias="transforms.MaskField.fields")
    transforms_MaskField_type: Optional[str] = Field(default=None, validation_alias="transforms.MaskField.type")
    flush_size: Optional[int] = Field(default=None, validation_alias="flush.size")
    rotate_interval_ms: Optional[int] = Field(default=None, validation_alias="rotate.interval.ms")
    rotate_schedule_interval_ms: Optional[int] = Field(default=None, validation_alias="rotate.schedule.interval.ms")
    time_partition_duration_ms: Optional[int] = Field(default=None, validation_alias="time.partition.duration.ms")
    time_path_format: Optional[str] = Field(default=None, validation_alias="time.path.format")
    filename_template: Optional[str] = Field(default=None, validation_alias="filename.template")
    compression: Optional[Literal["none","gz","snappy"]] = None
    errors_tolerance: Optional[Literal["none","all"]] = Field(default=None, validation_alias="errors.tolerance")
    errors_log_enable: Optional[Literal["true","false"]] = Field(default=None, validation_alias="errors.log.enable")
    errors_dlq_headers_enable: Optional[Literal["true","false"]] = Field(default=None, validation_alias="errors.deadletterqueue.context.headers.enable")

    @field_validator("timezone")
    @classmethod
    def tz_valid(cls, v):
        if v is None or v == "":
            return v
        if v not in pytz.all_timezones:
            raise ValueError(f"Invalid timezone: {v}")
        return v

    @field_validator("transforms_MaskField_type")
    @classmethod
    def mask_type_looks_right(cls, v, info):
        transforms = info.data.get("transforms")
        fields = info.data.get("transforms_MaskField_fields")
        if transforms and "MaskField" in transforms:
            if not fields:
                raise ValueError("MaskField configured but transforms.MaskField.fields is blank")
            if not v or "$" not in v:
                raise ValueError("MaskField type should look like org.apache.kafka.connect.transforms.MaskField$Value")
        return v

    @field_validator("flush_size","rotate_interval_ms","rotate_schedule_interval_ms","time_partition_duration_ms")
    @classmethod
    def positive_ints(cls, v):
        if v is None:
            return v
        if int(v) <= 0:
            raise ValueError("Must be a positive integer")
        return v

class SourceExtras(BaseModel):
    topic_regex_list: Optional[str] = Field(default=None, validation_alias="topic-regex.list")
    topics: Optional[str] = None
    gcs_credentials_json: Optional[str] = Field(default=None, validation_alias="gcs.credentials.json")
    gcs_bucket_name: str = Field(validation_alias="gcs.bucket.name")
    input_data_format: AllowedFormats = Field(validation_alias="input.data.format")
    output_data_format: AllowedFormats = Field(validation_alias="output.data.format")
    value_converter_decimal_format: Optional[Literal["NUMERIC","STRING"]] = Field(default=None, validation_alias="value.converter.decimal.format")
    value_converter_replace_null: Optional[Literal["true","false"]] = Field(default=None, validation_alias="value.converter.replace.null.with.default")
    gcs_pattern: Optional[str] = Field(default=None, validation_alias="gcs.pattern")
    file_regex: Optional[str] = Field(default=None, validation_alias="file.regex")
    poll_interval_ms: Optional[int] = Field(default=None, validation_alias="poll.interval.ms")

    @field_validator("poll_interval_ms")
    @classmethod
    def poll_pos(cls, v):
        if v is None:
            return v
        if int(v) <= 0:
            raise ValueError("poll.interval.ms must be positive")
        return v

    @field_validator("topics")
    @classmethod
    def topics_or_regex(cls, v, info):
        if (v is None or v == "") and not info.data.get("topic_regex_list"):
            raise ValueError("Provide either 'topics' or 'topic-regex.list'")
        return v

# -------------------
# IO & assembly
# -------------------

def load_excel_frames(path: str) -> Dict[str, pd.DataFrame]:
    xl = pd.ExcelFile(path)
    frames: Dict[str, pd.DataFrame] = {}
    for sheet in ["Common", "GCS_Sink", "GCS_Source"]:
        if sheet in xl.sheet_names:
            df = xl.parse(sheet).fillna("")
            df.columns = [normalize_key(str(c)) for c in df.columns]
            frames[sheet] = df
        else:
            frames[sheet] = pd.DataFrame()
    return frames

def build_rows(frames: Dict[str, pd.DataFrame]) -> List[Dict[str, Any]]:
    common = frames.get("Common", pd.DataFrame())
    sink = frames.get("GCS_Sink", pd.DataFrame())
    source = frames.get("GCS_Source", pd.DataFrame())

    rows: List[Dict[str, Any]] = []
    for _, c in common.iterrows():
        connector = {normalize_key(k): str(v).strip() for k, v in c.items() if str(v).strip() != ""}
        ctype = connector.get("connector_type", "")
        name = connector.get("name", "")
        if not ctype or not name:
            continue

        if ctype == "sink":
            tdf = sink[sink["name"] == name] if "name" in sink.columns else pd.DataFrame()
        else:
            tdf = source[source["name"] == name] if "name" in source.columns else pd.DataFrame()

        extra: Dict[str, Any] = {}
        if not tdf.empty:
            extra = {normalize_key(k): str(v).strip() for k, v in tdf.iloc[0].items() if str(v).strip() != ""}

        combined = {**connector, **extra}
        # Coerce boolean-like values to 'true'/'false'
        for _k in list(combined.keys()):
            if _k in BOOLISH_KEYS:
                combined[_k] = _coerce_boolish(combined[_k])

        combined.setdefault("cloud.provider", "gcp")
        combined.setdefault("kafka.auth.mode", "SERVICE_ACCOUNT")
        if ctype == "sink":
            combined.setdefault("connector.class", "GcsSink")
        else:
            combined.setdefault("connector.class", "GcsSource")
        combined.setdefault("tasks.max", "1" if ctype == "sink" else "5")
        combined.setdefault("status", "PAUSED")
        rows.append(combined)
    return rows

def validate_row(row: Dict[str, Any]) -> List[str]:
    errors: List[str] = []
    base_req = ["name", "cloud.environment", "kafka.endpoint", "kafka.region",
                "kafka.service.account.id", "topics.dir", "tasks.max"]
    for k in base_req:
        if not row.get(k):
            errors.append(f"Missing required field: {k}")

    if row.get("connector_type") == "sink":
        for k in ["topics", "gcs.bucket.name", "input.data.format", "output.data.format"]:
            if not row.get(k):
                errors.append(f"[sink] Missing required field: {k}")
    elif row.get("connector_type") == "source":
        if not (row.get("topic-regex.list") or row.get("topics")):
            errors.append("[source] Provide either 'topic-regex.list' or 'topics'")
        for k in ["gcs.bucket.name", "input.data.format", "output.data.format"]:
            if not row.get(k):
                errors.append(f"[source] Missing required field: {k}")
    else:
        errors.append("connector_type must be 'sink' or 'source'")
    return errors

def pydantic_validate(row: Dict[str, Any]) -> List[str]:
    msgs: List[str] = []
    try:
        _ = CommonRow.model_validate(row)
    except ValidationError as ve:
        msgs.extend([f"Common: {e['msg']} (at {'.'.join([str(loc) for loc in e['loc']])})" for e in ve.errors()])

    if row.get("connector_type") == "sink":
        try:
            _ = SinkExtras.model_validate(row)
        except ValidationError as ve:
            msgs.extend([f"Sink: {e['msg']} (at {'.'.join([str(loc) for loc in e['loc']])})" for e in ve.errors()])
    elif row.get("connector_type") == "source":
        try:
            _ = SourceExtras.model_validate(row)
        except ValidationError as ve:
            msgs.extend([f"Source: {e['msg']} (at {'.'.join([str(loc) for loc in e['loc']])})" for e in ve.errors()])

    return msgs

def assemble_config(row: Dict[str, Any]) -> Dict[str, Any]:
    control_keys = {"connector_type", "status"}
    config = {normalize_key(k): v for k, v in row.items() if k not in control_keys}
    return config

def write_json(path: Path, obj: Any):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(obj, f, indent=2, ensure_ascii=False)
        f.write("\n")

# -------------------
# Template creation
# -------------------
def make_template(path: str) -> None:
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # Common
    ws = wb.active
    ws.title = "Common"
    common_headers = [
        "connector_type","name","cloud.environment","kafka.endpoint","kafka.region",
        "kafka.service.account.id","topics.dir","tasks.max","status"
    ]
    ws.append(common_headers)
    ws.append(["sink","GcsSinkConnector_sample","prod","SASL_SSL://abc:9092","us-east4","sa-xxxxx","sink","1","RUNNING"])
    ws.append(["source","GcsSourceConnector_sample","prod","SASL_SSL://abc:9092","us-east4","sa-yyyyy","file/splitfiles_sample","5","PAUSED"])

    dv_type = DataValidation(type="list", formula1='"sink,source"')
    ws.add_data_validation(dv_type); dv_type.add("A2:A1048576")
    dv_status = DataValidation(type="list", formula1='"RUNNING,PAUSED"')
    ws.add_data_validation(dv_status); dv_status.add("I2:I1048576")
    dv_env = DataValidation(type="list", formula1='"dev,test,stage,prod"')
    ws.add_data_validation(dv_env); dv_env.add("C2:C1048576")
    for idx, _ in enumerate(common_headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = 28

    # GCS_Sink
    ws2 = wb.create_sheet("GCS_Sink")
    sink_headers = [
        "name","topics","gcs.credentials.config","gcs.bucket.name","input.data.format","output.data.format",
        "time.interval","timezone","transforms","transforms.MaskField.fields","transforms.MaskField.type",
        "flush.size","rotate.interval.ms","rotate.schedule.interval.ms","time.partition.duration.ms",
        "time.path.format","filename.template","compression","errors.tolerance","errors.log.enable",
        "errors.deadletterqueue.context.headers.enable"
    ]
    ws2.append(sink_headers)
    ws2.append([
        "GcsSinkConnector_sample","t1,t2","secret://gcs/sink/creds","bkt-d-use4-gcs-landing","JSON","JSON","HOURLY","US/Eastern",
        "MaskField","SSN","org.apache.kafka.connect.transforms.MaskField$Value",
        "100000","900000","","3600000","'year'=YYYY/'month'=MM/'day'=dd/'hour'=HH",
        "${topic}-${timestamp}.json","gz","all","true","true"
    ])
    for idx, _ in enumerate(sink_headers, start=1):
        ws2.column_dimensions[get_column_letter(idx)].width = 36
    dv_formats = DataValidation(type="list", formula1='"JSON,AVRO,PARQUET,CSV,STRING"')
    ws2.add_data_validation(dv_formats); dv_formats.add("E2:E1048576"); dv_formats.add("F2:F1048576")
    dv_time = DataValidation(type="list", formula1='"MINUTE,HOURLY,DAILY"')
    ws2.add_data_validation(dv_time); dv_time.add("G2:G1048576")
    dv_comp = DataValidation(type="list", formula1='"none,gz,snappy"')
    ws2.add_data_validation(dv_comp); dv_comp.add("R2:R1048576")
    dv_tol = DataValidation(type="list", formula1='"none,all"')
    ws2.add_data_validation(dv_tol); dv_tol.add("S2:S1048576")
    dv_bool = DataValidation(type="list", formula1='"true,false"')
    ws2.add_data_validation(dv_bool); dv_bool.add("T2:T1048576"); ws2.add_data_validation(dv_bool); dv_bool.add("U2:U1048576")

    # GCS_Source
    ws3 = wb.create_sheet("GCS_Source")
    source_headers = [
        "name","topic-regex.list","topics","gcs.credentials.json","gcs.bucket.name",
        "input.data.format","output.data.format","value.converter.decimal.format",
        "value.converter.replace.null.with.default","gcs.pattern","file.regex","poll.interval.ms"
    ]
    ws3.append(source_headers)
    ws3.append([
        "GcsSourceConnector_sample","DATA_**","","secret://gcs/source/creds","bkt-a-use4-gcs-landing","STRING","STRING",
        "NUMERIC","false","","","60000"
    ])
    for idx, _ in enumerate(source_headers, start=1):
        ws3.column_dimensions[get_column_letter(idx)].width = 36
    # add dropdown for the boolean field in source sheet
    dv_bool_src = DataValidation(type="list", formula1='"true,false"')
    ws3.add_data_validation(dv_bool_src); dv_bool_src.add("I2:I1048576")

    notes = wb.create_sheet("Notes")
    notes["A1"] = (
        "Instructions:\n"
        "1) Fill Common (one row per connector).\n"
        "2) Fill GCS_Sink or GCS_Source with a row that matches the same 'name'.\n"
        "3) Use secret refs (e.g., secret://...) for credentials.\n"
        "4) Boolean-like fields accept Yes/No, True/False, 1/0 and will be normalized to 'true'/'false'.\n"
        "5) Run: python gcs_config_generator.py generate --excel <file> --outdir out/configs\n"
    )

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)

# -------------------
# Generate CONFIG JSONs
# -------------------
def generate_configs(excel: str, outdir: str) -> None:
    frames = load_excel_frames(excel)
    rows = build_rows(frames)

    all_errs: List[str] = []
    for r in rows:
        errs = validate_row(r) + pydantic_validate(r)
        if errs:
            all_errs += [f"[{r.get('name','<unknown>')}] {m}" for m in errs]

    if all_errs:
        print("Validation failed:\n" + "\n".join(all_errs))
        raise SystemExit(1)

    out = Path(outdir)
    out.mkdir(parents=True, exist_ok=True)
    for r in rows:
        cfg = assemble_config(r)
        path = out / f"{r['name']}.config.json"
        write_json(path, cfg)
        print(f"[OK] wrote {path}")

# -------------------
# CLI
# -------------------
def main():
    ap = argparse.ArgumentParser(description="GCS Connector Config Generator (Excel -> JSON)")
    sub = ap.add_subparsers(dest="cmd", required=True)

    ap_t = sub.add_parser("template", help="Create Excel template")
    ap_t.add_argument("--out", default="templates/gcs_connectors_template.xlsx")

    ap_g = sub.add_parser("generate", help="Generate config JSONs from Excel")
    ap_g.add_argument("--excel", required=True)
    ap_g.add_argument("--outdir", default="out/configs")

    args = ap.parse_args()

    if args.cmd == "template":
        make_template(args.out)
        print(f"[OK] wrote {args.out}")
    elif args.cmd == "generate":
        generate_configs(args.excel, args.outdir)

if __name__ == "__main__":
    main()
